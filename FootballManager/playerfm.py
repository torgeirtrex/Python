import random
import pandas
from openpyxl import load_workbook

class Player:
    '''
    Player is on a single team with many other players
    Players play a in a game for a team
    '''
    def __init__(self, name, skill):
        self.name = name

        # Skill
        self.skill = skill
        # Salary

    def salary(self):
        return 5000 + self.skill * 10

    def __str__(self):
        return 'Name: {}, Skill: {}'.format(self.name, self.skill)


def generate_player():
    #Import player names from external file
#    df = pandas.read_excel('PlayersFMimport.xlsx', sheetname=0)
#   first_names2 = df['A'].tolist()
    wb = load_workbook('PlayersFMimport.xlsx')
    ws = wb.get_sheet_by_name('Ark1')
    column = ws['D']
    playernames = [column[x].value for x in range(len(column))]

    player_name =  random.choice(playernames)
    'player_name = random.shuffle(playernames)'
    'last_name = random.choice(last_names)'

    full_name = '{}'.format(player_name)
    skill = 10 + random.randint(0, 90)
    return Player(full_name, skill)
