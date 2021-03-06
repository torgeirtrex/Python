import random
from helpers import goals_scored
from openpyxl import load_workbook

class Team:
    '''
    player
    ranking
    single mananger
    '''
    pass

    def __init__(self, name):
        self.name = name
        self.players = []

        self.wins = 0
        self.losses = 0
        self.draws = 0
        self.total_points = 0
        'self.goalsseasonhometeam = self.game.play()[0]'
        'self.goalsagainstseason = 0'

        self.money = 1000000

    def pay_players(self):
        for player in self.players:
            self.money -= player.salary()

    def rating(self):
        "What is the rating of the team"
        rating = 0
        for player in self.players:
            rating += player.skill

        return rating

    def __str__(self):
        return '{} {}'.format(self.name, self.rating())


def generate_teams():
        # Import player names from external file

        wb = load_workbook('TeamImportFM.xlsx')
        ws = wb.get_sheet_by_name('Ark1')
        column = ws['A']
        team_names = [column[x].value for x in range(len(column))]

        team_name = random.choice(team_names)

        teamname = '{}'.format(team_name)
        return Team(teamname)


class Game:
    '''
    plays a game between two teams
    game belongs to a leagues
    '''
    def __init__(self, league, home_team, away_team):
        self.league = league
        self.home_team = home_team
        self.away_team = away_team

        self.home_team_won = None
        self.home_team_draw = None
        'self.goalsawayteam = 0'

        print('{} vs. {}'.format(self.home_team, self.away_team))

    def play(self):
        #Return who won
        #True means hometeam won, false away team win
        print('Play begins')
        #goals_by_home_team = goals_scored(self.home_team.rating())
        #goals_by_away_team = goals_scored(self.away_team.rating())
        goalshometeam = self.goals_scored(self.home_team.rating())
        goalsawayteam = self.goals_scored(self.away_team.rating())

        print('{}'.format(self.home_team), goalshometeam, '-', '{}'.format(self.away_team), goalsawayteam)
        '''print(f"Home team scored {goalshometeam}")'''

        'self.goalsseasonhometeam += goalshometeam'
        if goalshometeam == goalsawayteam:
            print('Match is a draw')
            self.home_team_draw = True
        elif goalshometeam > goalsawayteam:
            print('{} wins'.format(self.home_team))
            self.home_team_won = True
            self.home_team_goals = goalshometeam
        else:
            print('{} wins.'.format(self.away_team))
            self.home_team_won = False
        print('Play ends')
        return goalshometeam, goalsawayteam


    def countgoalshome(self):
        goalcounterhome = self.play()[0]
        return goalcounterhome

    def countgoalsaway(self):
        goalcounteraway = self.play()[1]
        return goalcounteraway

#Function to let rating influence goals scored but add an element of luck to goalscoring
    def goals_scored(self, rating):
        # rating * random number between 1 to 5 / rating
        return int((rating * random.randint(0, 5) / rating))
#TO DO
    def print_menu(self):
        print('1. Select to view your squad')
        print('2. Do something else')





